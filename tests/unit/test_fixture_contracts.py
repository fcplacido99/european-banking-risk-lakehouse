from csv import DictReader
from decimal import Decimal, InvalidOperation
from pathlib import Path

import pytest
from openpyxl import load_workbook


FIXTURE_ROOT = Path(__file__).parents[1] / "fixtures"
CSV_ROOT = FIXTURE_ROOT / "csv"
XLSX_ROOT = FIXTURE_ROOT / "xlsx"

CREDIT_HEADER = [
    "LEI_Code",
    "NSA",
    "Period",
    "Item",
    "Label",
    "Portfolio",
    "Country",
    "Country_rank",
    "Exposure",
    "Status",
    "Perf_Status",
    "NACE_codes",
    "Amount",
    "Footnote",
    "Row",
    "Column",
    "Sheet",
]

OTHER_HEADER = [
    "LEI_Code",
    "NSA",
    "Period",
    "Item",
    "Label",
    "ASSETS_FV",
    "ASSETS_Stages",
    "Exposure",
    "Financial_instruments",
    "Amount",
    "Fin_end_year",
    "n_quarters",
    "Footnote",
    "Row",
    "Column",
    "Sheet",
]

CREDIT_NATURAL_KEY = [
    "LEI_Code",
    "NSA",
    "Period",
    "Item",
    "Portfolio",
    "Country",
    "Country_rank",
    "Exposure",
    "Status",
    "Perf_Status",
    "NACE_codes",
]


def read_csv_fixture(
    filename: str,
) -> tuple[list[str], list[dict[str, str]]]:
    """Return the header and rows from one UTF-8 CSV fixture."""

    path = CSV_ROOT / filename

    with path.open(encoding="utf-8", newline="") as source:
        reader = DictReader(source)
        rows = list(reader)

    assert reader.fieldnames is not None
    return reader.fieldnames, rows


@pytest.mark.parametrize(
    ("filename", "expected_header", "expected_rows"),
    [
        ("tr_cre_valid.csv", CREDIT_HEADER, 1),
        ("tr_oth_valid.csv", OTHER_HEADER, 1),
        ("tr_cre_duplicate_natural_key.csv", CREDIT_HEADER, 2),
        ("tr_oth_nonnumeric_amount.csv", OTHER_HEADER, 1),
        ("npe_zero_denominator.csv", CREDIT_HEADER, 2),
    ],
)
def test_fact_fixture_headers_and_row_counts(
    filename: str,
    expected_header: list[str],
    expected_rows: int,
) -> None:
    header, rows = read_csv_fixture(filename)

    assert header == expected_header
    assert len(rows) == expected_rows


def test_valid_credit_fixture_contains_expected_edge_cases() -> None:
    _, rows = read_csv_fixture("tr_cre_valid.csv")
    row = rows[0]

    assert row["LEI_Code"] == "0w2pzjm8xoy22m4gg883"
    assert row["Country"] == "0"
    assert row["Country_rank"] == "0"
    assert row["Status"] == "0"
    assert row["Perf_Status"] == "0"
    assert row["NACE_codes"] == "0"
    assert row["Footnote"] == ""


def test_credit_duplicate_fixture_repeats_the_natural_key() -> None:
    _, rows = read_csv_fixture("tr_cre_duplicate_natural_key.csv")

    keys = {
        tuple(row[column] for column in CREDIT_NATURAL_KEY)
        for row in rows
    }

    assert len(keys) == 1
    assert rows[0]["Amount"] != rows[1]["Amount"]
    assert rows[0]["Row"] != rows[1]["Row"]


def test_nonnumeric_amount_fixture_cannot_be_parsed_as_decimal() -> None:
    _, rows = read_csv_fixture("tr_oth_nonnumeric_amount.csv")

    with pytest.raises(InvalidOperation):
        Decimal(rows[0]["Amount"])


def test_zero_denominator_fixture_preserves_source_values() -> None:
    _, rows = read_csv_fixture("npe_zero_denominator.csv")
    amounts_by_status = {
        row["Perf_Status"]: Decimal(row["Amount"])
        for row in rows
    }

    assert amounts_by_status["0"] == Decimal("0")
    assert amounts_by_status["2"] == Decimal("10")


def test_reused_item_fixture_requires_template_in_join_key() -> None:
    header, rows = read_csv_fixture("sdd_reused_item.csv")

    assert header == ["CSV", "Template", "Item", "Category", "Label"]
    assert {row["Item"] for row in rows} == {"2420502"}
    assert {row["Template"] for row in rows} == {
        "Credit Risk_IRB_a",
        "Credit Risk_IRB_b",
        "Credit Risk_STA_a",
        "Credit Risk_STA_b",
    }


def test_workbook_fixture_has_two_banks_but_later_physical_rows() -> None:
    path = XLSX_ROOT / "metadata_trailing_formatted_blanks.xlsx"
    workbook = load_workbook(path, read_only=False, data_only=True)
    sheet = workbook["Other banks"]

    header = [
        sheet.cell(row=3, column=column).value
        for column in range(1, 5)
    ]
    bank_rows = [
        row_number
        for row_number in range(4, sheet.max_row + 1)
        if sheet.cell(row=row_number, column=3).value not in (None, "")
    ]

    assert header == ["Country", "Desc_country", "LEI_Code", "Name"]
    assert sheet.max_row == 50
    assert bank_rows == [4, 5]
    assert sheet.cell(row=6, column=1).value is not None
    assert sheet.cell(row=6, column=3).value is None
    assert sheet.cell(row=50, column=1).value is None
    assert sheet.cell(row=50, column=1).fill.fill_type == "solid"

    workbook.close()


def test_combined_fixture_size_remains_repository_safe() -> None:
    total_bytes = sum(
        path.stat().st_size
        for path in FIXTURE_ROOT.rglob("*")
        if path.is_file()
    )

    assert total_bytes < 100_000